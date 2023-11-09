import openpyxl as pyxl

class ExcelFile:
    def __init__(self):
        self.wb = None
        self.sheet = None

    def create(self):
        '''create excel file in memory and return active sheet'''
        try:
            self.wb = pyxl.Workbook()
            self.sheet = self.wb.active
        except Exception as e:
            print(f"Error create_workbook(): {e}")
            raise RuntimeError

    def open(self, filepath):
        '''Open excel file, store in memory, return active sheet'''
        try:
            self.wb = pyxl.load_workbook(filepath)
            self.sheet = self.wb.active
        except Exception as e:
            print(f"Error load_workbook(): {e}")
            raise FileNotFoundError

    def save(self, filepath):
        '''Save as .xlsx file'''
        try:
            self.wb.save(filepath)
        except Exception as e:
            self.wb.save('backup.xlsx')
            print(f"Error save(): {e} \n Saved backup.xlsx")
            raise RuntimeError
        
    def read_column(self, column='A', start=1):
        '''Read first column'''
        list = []
        for cell in self.sheet[column][start:]:
            print(cell.value)
            list.append(cell.value)
        return list

    def write_cell(self, value, row, column):
        '''Write cell'''
        self.sheet.cell(row=row, column=column).value = value

    def write_row(self, row: list, index = 1):
        '''(Over)write first row, aka column names'''
        for i, value in enumerate(row, start=1):
            self.sheet.cell(row=index, column=i).value = value

    def append_rows(self, rows):
        '''Append rows '''
        for row in rows:
            self.sheet.append(row)

    def close(self):
        self.wb.close()